from operator import itemgetter
import os
import re

from flask import current_app
from openpyxl import load_workbook
from PyPDF2 import PdfReader


def parse_deluxe_invoice(directory):
    """Parse a Deluxe invoice an get key information.

    Key info:
    - Invoice date
    - Invoice number
    - Film title
    - Delivery date
    """
    invoices = []
    for invoice in os.listdir(directory):
        invoice_data = {}

        new_path = os.path.join(directory, invoice)

        pdf = PdfReader(new_path)
        pdf_text = pdf.pages[0].extract_text()

        # get invoice date
        invoice_date = re.search(
            '(?<=Due Date:)(.*)(?=Invoice Number)(?s)', pdf_text
        ).group(0).strip()
        invoice_data['Invoice Date'] = invoice_date

        # get invoice number
        invoice_number = re.search(
            '(?<=Invoice Date:)(.*)(?=Customer Account No:)(?s)', pdf_text
        ).group(0).strip()
        invoice_data['Invoice Number'] = invoice_number

        # get film title
        film_title = re.search(
            '(?<=Title: )(.*)', pdf_text
        ).group(0).strip()
        invoice_data['Film'] = film_title.upper()

        # get open date
        open_date = re.search(
            '(?<=OPEN DATE: )(.*)(?=Pay Online)(?s)', pdf_text
        ).group(0).strip()
        invoice_data['Open Date'] = open_date

        # get fee
        fee = re.search('(?<=TAX AMOUNT:)(.*)(?=  GRAND TOTAL:)(?s)', pdf_text
                        ).group(0).strip()
        invoice_data['Fee'] = float(fee[1:])

        invoices.append(invoice_data)
    return invoices


def create_po(invoices):
    """Create the PO."""
    # Cell D31 - {film title}: Delivery Fee - {date} Invoice # {invoice number}
    directory = current_app.config['PO_FOLDER']
    purchase_order = os.listdir(directory)[0]

    wb = load_workbook(filename=f'{directory}/{purchase_order}')
    ws = wb['PO Form']

    cell_num = 31
    count = 0
    sorted_invoices = sorted(invoices, key=itemgetter('Invoice Number'))
    for invoice in sorted_invoices:
        ws['B29'] = invoice['Invoice Date']
        ws[f'D{cell_num + count}'] = (f"{invoice['Film']}: Delivery Fee - "
                                      f"{invoice['Open Date']} "
                                      f"Invoice # {invoice['Invoice Number']}"
                                      )
        ws[f'V{cell_num + count}'] = invoice['Fee']

        count += 1

    wb.save(filename=f'{directory}/{purchase_order}')
