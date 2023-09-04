import os

from docx import Document
from flask import current_app
from openpyxl import load_workbook
import tmdbsimple as tmdb

tmdb.API_KEY = os.getenv('TMDB_API_KEY')
tmdb.REQUESTS_TIMEOUT = 5


class Film:
    """A class to represent a film and its AFI-style credits.

    The only attributes created when the class is instantiated are the
    title and imdb_id, both of which are read from a spreadsheet.

    All other attributes are created by calling class methods.

    Attributes:
        title (str): the name of the film
        imdb_id (str): the unique IMDb ID for the film, e.g. tt0094625

        tmdb_id (str):  the unique TMDB (TheMoveDatabase) ID for the film
        movie_info (str): a TMDB JSON response with the film's primary info
        directors (lst): a list of the film's directors
        writers (lst): a list of the film's writers
        producers (lst): a list of the film's producers
        countries (lst): a list of the film's production countries
        languages (lst): a list of the film's spoken languages
        release_date (str): the film's year of release
        runtime (str): the film's runtime
        rating (str): the film's MPAA rating
        genres (lst): a list of the film's genres
    """

    def __init__(self, title, imdb_id):
        """Initialize the instance with the film's title and IMDB ID."""
        self.title = title
        self.imdb_id = imdb_id
        self.tmdb_id = None
        self.movie_info = None
        self.directors = []
        self.writers = []
        self.producers = []
        self.countries = []
        self.languages = []
        self.release_date = None
        self.runtime = None
        self.rating = None
        self.genres = []
        self.trailer = None

    def get_tmdb_id(self):
        """Get the film's TMDB ID and assigns it to the tmdb_id attribute."""
        response = tmdb.Find(self.imdb_id).info(external_source='imdb_id')
        self.tmdb_id = response['movie_results'][0]['id']

    def get_movie_info(self):
        """Get the film's primary information."""
        self.movie_info = tmdb.Movies(self.tmdb_id).info()

    def get_release_date(self):
        """Get the film's release year.

        TMDB's release_date field includes the month and day which we slice.
        """
        self.release_date = self.movie_info['release_date'][:4]

    def get_runtime(self):
        """Get the film's runtime."""
        self.runtime = self.movie_info['runtime']

    def get_countries(self):
        """Get the film's production countries."""
        production_countries = self.movie_info['production_countries']
        countries = []
        for country in production_countries:
            # the production_countries dictionary uses the two-letter
            # ISO 3166-1 code
            if country['iso_3166_1'] == 'US':
                # TMDB style writes out "United States of America" so we alter
                # this to match AFI style
                countries.append('U.S.')
            else:
                countries.append(country['name'])
        self.countries = countries

    def get_languages(self):
        """Get the film's spoken languages."""
        spoken_languages = self.movie_info['spoken_languages']
        languages = []
        for language in spoken_languages:
            # must use english_name because the name key returns the language's
            # name IN the language
            languages.append(language['english_name'])
        self.languages = languages

    def get_crew(self):
        """Get the film's director(s), screenwriter(s), and producer(s)."""
        crew = tmdb.Movies(self.tmdb_id).credits()['crew']

        for crew_member in crew:
            if crew_member['job'] == 'Director':
                self.directors.append(crew_member['name'])
            if (
                # A film's writers can be credited on TMDB as either 'Writer'
                # or 'Screenplay' so account for both options
                crew_member['job'] == 'Screenplay' or
                crew_member['job'] == 'Writer'
            ):
                self.writers.append(crew_member['name'])
            # this should exclude Executive Producer and Associate Producer
            if crew_member['job'] == 'Producer':
                self.producers.append(crew_member['name'])

    def get_rating(self):
        """Get the film's MPAA rating aka certification."""
        response = tmdb.Movies(self.tmdb_id).releases()
        for country in response['countries']:
            # we don't need the certification for any other country
            if country['iso_3166_1'] == 'US':
                if country['certification'] == '':
                    # a blank rating implies it is NOT RATED, per AFI style
                    self.rating = 'NOT RATED'
                else:
                    self.rating = f'RATED {country["certification"]}'

    def get_genres(self):
        """Get the film's genres."""
        tmdb_genres = self.movie_info['genres']
        genres = []
        for genre in tmdb_genres:
            genres.append(genre['name'])
        self.genres = genres

    def get_trailer(self):
        """Get the film's videos."""
        video_list = tmdb.Movies(self.tmdb_id).videos()['results']
        # older films likely won't have trailers
        if video_list is False:
            self.trailer = ''
        else:
            for video in video_list:
                # the Video API call returns various types of videos, including
                # teasers, clips, featurettes, etc. but we only want a trailer
                if video['type'] == 'Trailer':
                    trailer = f'https://www.youtube.com/watch?v={video["key"]}'
                    self.trailer = trailer
                    # a film can have multiple trailers and there is no way to
                    # programatically tell which is the best one to use so just
                    # grab the first one then exit the loop
                    break


def get_credits(directory):
    """Loop over the uploaded workbook and get the credits for each film."""
    directory = current_app.config['CREDITS_FOLDER']
    manifest = os.listdir(directory)[0]

    wb = load_workbook(filename=f'{directory}/{manifest}')
    ws = wb.active

    # Python's range functiion excludes the second parameter so we add 1 so it
    # can be included, otherwise it will skip whatever film isi n the last row
    last_row = ws.max_row + 1
    first_row = 2

    films = []

    for row in range(first_row, last_row):
        film = Film(title=ws.cell(row, 1).value,
                    imdb_id=ws.cell(row, 2).value)
        films.append(film)

    # create a blank Word doc
    document = Document()

    for film in films:
        # each class attribute is created by calling its associated method
        film.get_tmdb_id()
        film.get_movie_info()
        film.get_release_date()
        film.get_runtime()
        film.get_countries()
        film.get_languages()
        film.get_crew()
        film.get_rating()

        # if a film is solely in English, it is not explicitly noted
        if (len(film.languages) == 1) and (film.languages[0] == 'English'):
            language = ''
        else:
            language = f'In {"/".join(film.languages)} with English subtitles. '

        paragraph = (
            f'DIR {", ".join(film.directors)}; '
            f'SCR {", ".join(film.writers)}; '
            f'PROD {", ".join(film.producers)}. '
            f'{"/".join(film.countries)}, '
            f'{film.release_date}, color/b&w, {film.runtime} min. '
            f'{language}'
            f'{film.rating}'
        )

        document.add_paragraph(paragraph)

    # os.path.join accounts for OS-dependent path separators
    document.save(os.path.join(
        current_app.config['CREDITS_FOLDER'],
        'credits.docx'))
