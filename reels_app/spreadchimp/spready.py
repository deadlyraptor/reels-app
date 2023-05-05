import os
import csv

from flask import current_app
from openpyxl import load_workbook


def spready(directory):
    """Compile the different series into a list."""
    directory = current_app.config['SPREADSHEET_FOLDER']
    manifest = os.listdir(directory)[0]

    wb = load_workbook(filename=f'{directory}/{manifest}')
    ws = wb.active

    last_row = ws.max_row
    first_row = 7

    series = []
    for row in range(first_row, last_row):
        series_title = ws.cell(row, 23).value
        if series_title in series or series_title is None:
            continue
        else:
            series.append(series_title)

    def prep_contacts(series):
        """Compile contacts into a list.

        Arguments:
            series = The series the customer attended. Note that the original
            workbook has the film name but users change it to the series before
            running the workbook through the script.

        Returns:
            contacts = The list of all contacts for a given series.
        """
        contacts = []
        for row in range(first_row, last_row):
            contact = []
            opt_in = ws.cell(row, 8).value  # email opt-in, col H
            series_title = ws.cell(row, 23).value  # event name, col W
            if not opt_in:
                continue
            elif opt_in and series_title == series:
                contact = [ws.cell(row, 7).value,  # email, col G
                           ws.cell(row, 3).value.title(),  # first name, col C
                           ws.cell(row, 5).value.title(),  # last name, col E
                           ws.cell(row, 19).value.replace
                           ('No Primary Phone', ''),  # phone number, col S
                           ws.cell(row, 14).value,  # address 1, col N
                           ws.cell(row, 15).value,  # address 2, col O
                           ws.cell(row, 16).value.title(),  # city, col P
                           ws.cell(row, 17).value,  # state, col Q
                           ws.cell(row, 18).value,  # postal code, col R
                           ]
                contacts.append(contact)
        return contacts

    headers = ['Email', 'First Name', 'Last Name', 'Phone',
               'Street Address', 'Address Line 2', 'City', 'State/Prov/Region',
               'Zip/Postal']

    for series in series:
        contacts = prep_contacts(series)
        with open(f'csvs/{series}.csv', mode='w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=',', quotechar='"',
                                quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)
            for contact in contacts:
                writer.writerow(contact)
