import os

from flask import current_app
from openpyxl import load_workbook
import tmdbsimple as tmdb

from reels_app.credit.utils import Film

tmdb.API_KEY = os.getenv('TMDB_API_KEY')
tmdb.REQUESTS_TIMEOUT = 5


def get_genres(directory):
    """Loop over the uploaded workbook and get the film's genres."""
    directory = current_app.config['GENRE_FOLDER']
    manifest = os.listdir(directory)[0]

    wb = load_workbook(filename=f'{directory}/{manifest}')
    ws = wb.active

    # Python's range functiion excludes the second parameter so we add 1 so it
    # can be included, otherwise it will skip whatever film isi n the last row

    last_row = ws.max_row + 1
    first_row = 2

    films = []

    for row in range(first_row, last_row):
        # this accounts for rows that are empty
        if ws.cell(row, 1).value == None:
            pass
        else:
            film = Film(title=ws.cell(row, 1).value,
                        imdb_id=ws.cell(row, 2).value)
            films.append(film)

    count = 0
    for film in films:
        film.get_tmdb_id()
        film.get_movie_info()
        film.get_genres()

        ws[f'G{first_row + count}'] = (', '.join(film.genres))

        count += 1

    wb.save(filename=f'{directory}/{manifest}')
